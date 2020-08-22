# -*- coding: utf-8 -*-
"""
Created on Sun Jul  5 12:53:34 2020

This file covers various functionalities of openpyxl libraries.
It would help the user to understand about how excel can be controlled/
modified using python. Following are majorly covered topics:
    1. INSTALLATION OF OPENPYXL AND BASICS
    2. OPENPYXL WRITE DATA TO CELL
    3. OPENPYXL APPEND VALUES
    4. OPENPYXL READ DATA FROM CELL
    5. OPENPYXL SHEETS
    6. OPENPYXL MERGING CELL AND FREEZE THE PANES
    7. OPENPYXL FORMULAS
    8. ADDING CHART TO EXCEL FILE
    9. ADDING IMAGE TO EXCEL

@author: Shehzaan_Sheikh
"""
# INSTALLATION OF OPENPYXL AND BASICS
#pip install openpyxl

from openpyxl import Workbook  
import time  
  
wb = Workbook()  
sheet = wb.active  
  
sheet['A1'] = 87  
sheet['A2'] = "Shehzaan"  
sheet['A3'] = 41.80  
sheet['A4'] = 10  
  
now = time.strftime("%x")  
sheet['A5'] = now  
  
wb.save("sample_file.xlsx")

#%%
# OPENPYXL WRITE DATA TO CELL
from openpyxl import load_workbook  
wb = load_workbook(r'.\sample_file.xlsx')  
  
sheet = wb.active  
sheet['A1'] = 'CRISIL'  
  
sheet.cell(row=2, column=2).value = 5  
wb.save(r'.\sample_file.xlsx')

#%%
# OPENPYXL APPEND VALUES

from openpyxl import Workbook  
  
wb = Workbook()  
sheet = wb.active  
  
data = (  
    (11, 48, 50),  
    (81, 30, 82),  
    (20, 51, 72),  
    (21, 14, 60),  
    (28, 41, 49),  
    (74, 65, 53),  
    ("Peter", 'Andrew',45.63)  
)  
  
for i in data:  
    sheet.append(i)  
wb.save('appending_values.xlsx')

#%%
# OPENPYXL READ DATA FROM CELL

import openpyxl  
  
wb = openpyxl.load_workbook('sample_file.xlsx')  
  
sheet = wb.active  
  
x1 = sheet['A1']  
x2 = sheet['A2']  
#using cell() function  
x3 = sheet.cell(row=3, column=1)  
  
print("The first cell value:",x1.value)  
print("The second cell value:",x2.value)  
print("The third cell value:",x3.value) 


#%%
# OPENPYXL SHEETS

import openpyxl  
wb = openpyxl.load_workbook('Weekdays.xlsx')  
  
#Getting list of all sheet available in workbook  
print(wb.sheetnames)
  
# Returning object  
active_sheet = wb.active  
print(type(active_sheet))  
  
# Title of sheet  
sheet = wb['Monday']
print(sheet.title)

sheet['A1'] = 87  
sheet['A2'] = "Shehzaan"  
sheet['A3'] = 41.80  
sheet['A4'] = 10

wb.save('Weekdays.xlsx')

#%%
# OPENPYXL MERGING CELL AND FREEZE THE PANES

from openpyxl.styles import Alignment  
  
wb = Workbook()  
sheet = wb.active
sheet.title = "sheet_name_new"
  
sheet.merge_cells('A1:B2')
# We also have unmerged_cells() method to unmerge the cells back again.

# We can freeze the pane by specifying the index
sheet.freeze_panes = 'A4'

  
cell = sheet.cell(row=1, column=1)  
cell.value = 'Shehzaan Sheikh'  
cell.alignment = Alignment(horizontal='center', vertical='center')  
  
wb.save('merging.xlsx')

#%%
# OPENPYXL FORMULAS

from openpyxl import Workbook  
wb = Workbook()  
sheet = wb.active  
  
rows_count = (  
    (14, 27),  
    (22, 30),  
    (42, 92),  
    (51, 32),  
    (16, 60),  
    (63, 13)  
)  
  
for i in rows_count:  
    sheet.append(i)  
  
cell = sheet.cell(row=7, column=3)  
cell.value = "=SUM(A1:B6)"  
cell.font = cell.font.copy(bold=True)  
  
wb.save('formulas_book.xlsx')

#%%
# ADDING CHART TO EXCEL FILE

from openpyxl import Workbook  
from openpyxl.chart import BarChart, Reference  
  
wb = Workbook()  
sheet = wb.active  
  
# Let's create some sample student data  
rows = [  
    ["Serial_no", "Roll no", "Marks"],  
    [1, "0090011", 75],  
    [2, "0090012", 60],  
    [3, "0090013", 43],  
    [4, "0090014", 97],  
    [5, "0090015", 63],  
    [6, "0090016", 54],  
    [7, "0090017", 86],  
]  
  
for i in rows:  
    sheet.append(i)  
  
chart = BarChart()  
values = Reference(worksheet=sheet,  
                 min_row=1,  
                 max_row=8,  
                 min_col=2,  
                 max_col=3)  
  
chart.add_data(values, titles_from_data=True)  
sheet.add_chart(chart, "E2")  
  
wb.save("chart.xlsx") 

#%%
# ADDING IMAGE TO EXCEL

from openpyxl import Workbook
from openpyxl.drawing.image import Image  
  
# You would need to install additional library
# pip install pillow

# Let's use the hello_world spreadsheet since it has less data  
wb = Workbook()  
spreadsheet = wb.active
  
logo = Image(r".\image.jpg")  
  
# A bit of resizing to not fill the whole spreadsheet with the logo  
logo.height = 150  
logo.width = 150  
  
spreadsheet.add_image(logo, "A1")
wb.save(filename="hello_world_image.xlsx")
